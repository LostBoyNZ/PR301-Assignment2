# Graham
import sys
from openpyxl import load_workbook

try:
    from errors import ErrorHandler as Err
except NameError and ModuleNotFoundError and ImportError:
    print("Fatal Error - Errors.py not found.")
    sys.exit()
except Exception as e:
    print("Exception: {}".format(e))

try:
    from data_processor import DataProcessor as dp
except NameError and ModuleNotFoundError and ImportError:
    print(Err.get_error_message(404, "data_processor"))
    sys.exit()
except Exception as e:
    print(Err.get_error_message(901, e))

try:
    from log_file_handler import LogFileHandler as Lfh
except NameError and ModuleNotFoundError and ImportError:
    print(Err.get_error_message(404, "log_file_handler"))
    sys.exit()
except Exception as e:
    print(Err.get_error_message(901, e))


class DatabaseExcel(object):  # Graham

    row_names = ['emp_id', 'gender', 'age', 'sales',
                 'bmi', 'salary', 'birthday', 'valid']
    the_sheet = None

    @staticmethod
    def get_input(text):
        return input(text)

    def choose_sheet(self, wb):
        sheet_names = wb.sheetnames
        # default to the first sheet
        sheet = wb.worksheets[0]

        if len(sheet_names) > 1:
            print("Sheets:\n")
            i = 1
            for sheet in sheet_names:
                print(str(i) + ") " + sheet)
                i += 1

            while True:
                sheet_number = self.get_input("\nEnter the sheet number to read >>> ")
                try:
                    sheet_number -= 1
                    sheet = wb.worksheets[int(sheet_number)]
                    break
                except IndexError or KeyError or ValueError:
                    print(Err.get_error_message(102, sheet_number))
                except TypeError:
                    print(Err.get_error_message(106, "numbers"))

        return sheet

    def choose_file(self):
        while True:
            file_name = self.get_input("\nPlease enter the excel file to read >>> ")
            wb = None

            try:
                wb = load_workbook(file_name)
                break
            except FileNotFoundError:
                print(Err.get_error_message(201))

        return wb

    def set_sheet(self, sheet):
        self.the_sheet = sheet

    def get_sheet(self):
        return self.the_sheet

    @staticmethod
    def get_sheet_size(sheet):
        sheet_size = {'Col Count': sheet.max_column,
                      'Row Count': sheet.max_row}

        return sheet_size

    @staticmethod
    def append_log(data_to_log):
        Lfh.append_file('log.txt', data_to_log)

    def check_for_duplicate_key(self, key, all_keys):
        if key in all_keys:
            self.append_log("Duplicate Key" + str(key))

    def get_person(self, target_row, target_col, columns_in_row):
        sheet = self.get_sheet(self)
        data_row = []
        col_num = 0
        col_count = self.get_sheet_size(sheet)['Col Count']
        for column in range(0, col_count):
            output = sheet.cell(row=target_row, column=target_col).value
            data_row.append(str(output))

            columns_in_row[self.row_names[col_num]] = data_row[col_num]
            target_col = target_col + 1
            col_num = col_num + 1

        return columns_in_row

    def add_staff_member_details(self, all_staff, id_key, columns_in_row):
        # Skip the ID and Valid rows by starting at 1 and stopping before end
        for row in self.row_names[1:-1]:
            all_staff[id_key][row] = columns_in_row[row]

        all_staff[id_key]['valid'] = "0"

        return all_staff

    @staticmethod
    def get_key_from_row(sheet, row):
        possible_key = sheet.cell(row=row, column=1).value
        validated_key = dp.validate_key(str(possible_key))

        return validated_key

    def count_rows(self):
        sheet = self.get_sheet(self)

        return self.get_sheet_size(sheet)['Row Count']

    def extract_staff_from_sheet(self):
        target_col = 2
        target_row = 1
        columns_in_row = {}
        all_id_keys = []
        all_staff = {}
        dup_id_keys = []
        row_count = self.count_rows(self)

        for row in range(0, row_count):
            id_key = self.get_key_from_row(self.get_sheet(self), target_row)
            all_id_keys.append(id_key)

            if self.check_for_duplicate_key(self, id_key, all_id_keys):
                dup_id_keys.append(id_key)

            columns_in_row = self.get_person(self, target_row, target_col,
                                             columns_in_row)

            all_staff[id_key] = {}
            all_staff =\
                self.add_staff_member_details(self, all_staff, id_key,
                                              columns_in_row)

            target_col = 1
            target_row += 1

        return all_staff, dup_id_keys

    def get_data_from_excel(self, wb, switch):
        try:
            self.set_sheet(self, self.choose_sheet(self, wb))
        except TypeError:
            self.set_sheet(self, self.choose_sheet(wb))

        all_staff, dup_id_keys = self.extract_staff_from_sheet(self)
        dict_valid = dp.send_to_validate(all_staff, switch, dup_id_keys)

        return dict_valid
