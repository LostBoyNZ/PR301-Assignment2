import io
import sys
from database_excel import DatabaseExcel
from unittest.mock import patch
import unittest
from openpyxl import load_workbook


class TestDatabaseExcel(unittest.TestCase):

    def test_choose_file_loads_a_workbook_successfully(self):
        # Arrange
        user_input = ['testdata\\grahams_test_data2.xlsx']
        expected_type = type(load_workbook('testdata\\grahams_test_data2.xlsx'))

        # Act
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.choose_file(DatabaseExcel)

        # Assert
        self.assertTrue(type(result) == expected_type)

    def test_choose_file_loads_a_workbook_fails_when_file_not_found(self):
        # Arrange
        user_input = ['no file.xlsx', 'testdata\\grahams_test_data2.xlsx']
        expected_output = "File not found"

        # Act
        captured_output = io.StringIO()
        sys.stdout = captured_output
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.choose_file(DatabaseExcel)
        sys.stdout = sys.__stdout__

        # Check if the printed output includes the expected strings I'm looking for
        if expected_output in captured_output.getvalue():
            result = True
        else:
            result = False

        # Assert
        self.assertTrue(result)

    def test_create_connection_valid_entries_extract_data_from_sheet(self):
        # Arrange
        user_input = [2]
        wb = load_workbook('testdata\\grahams_test_data2.xlsx')
        expected_data = {'Empid': {'gender': 'Age', 'age': 'Sales', 'sales': 'BMI', 'bmi': 'Salary',
                                   'salary': 'Birthday', 'birthday': 'None', 'valid': '0'},
                         'A001': {'gender': 'M', 'age': '52', 'sales': '123', 'bmi': 'Overweight', 'salary': '50',
                                  'birthday': '1998-10-23 00:00:00', 'valid': '0'},
                         'Z005': {'gender': 'F', 'age': '18', 'sales': '624', 'bmi': 'Normal', 'salary': '850',
                                  'birthday': '25/04/1158', 'valid': '1'},
                         'None': {'gender': 'None', 'age': 'None', 'sales': 'None', 'bmi': 'None', 'salary': 'None',
                                  'birthday': 'None', 'valid': '0'}}

        # Act
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.create_connection(DatabaseExcel, wb, "")

        # Assert
        self.assertTrue(result == expected_data)

    def test_create_connection_valid_entries_and_duplicate_keys_extract_data_from_sheet(self):
        # Arrange
        user_input = [3]
        wb = load_workbook('testdata\\grahams_test_data2.xlsx')
        expected_data = {'Empid': {'gender': 'Age', 'age': 'Sales', 'sales': 'BMI', 'bmi': 'Salary',
                                   'salary': 'Birthday', 'birthday': 'None', 'valid': '0'},
                         'A001': {'gender': 'M', 'age': '52', 'sales': '123', 'bmi': 'Overweight', 'salary': '50',
                                  'birthday': '1998-10-23 00:00:00', 'valid': '0'},
                         'Z005': {'gender': 'M', 'age': '50', 'sales': '500', 'bmi': 'Normal', 'salary': '500',
                                  'birthday': '1968-01-01 00:00:00', 'valid': '0'}}

        # Act
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.create_connection(DatabaseExcel, wb, "")

        # Assert
        self.assertTrue(result == expected_data)

    def test_create_connection_index_errror_produces_error_before_trying_again(self):
        # Arrange
        user_input = [99, 2]
        wb = load_workbook('testdata\\grahams_test_data2.xlsx')
        expected_output = "Invalid Input, please try again"

        # Act
        captured_output = io.StringIO()
        sys.stdout = captured_output
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.create_connection(DatabaseExcel, wb, "")
        sys.stdout = sys.__stdout__

        # Check if the printed output includes the expected strings I'm looking for
        if expected_output in captured_output.getvalue():
            result = True
        else:
            result = False

        # Assert
        self.assertTrue(result)

    def test_create_connection_text_instead_of_number_produces_error_before_trying_again(self):
        # Arrange
        user_input = ['a', 2]
        wb = load_workbook('testdata\\grahams_test_data2.xlsx')
        expected_output = "Invalid input - Only accepts numbers"

        # Act
        captured_output = io.StringIO()
        sys.stdout = captured_output
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.create_connection(DatabaseExcel, wb, "")
        sys.stdout = sys.__stdout__

        # Check if the printed output includes the expected strings I'm looking for
        if expected_output in captured_output.getvalue():
            result = True
        else:
            result = False

        # Assert
        self.assertTrue(result)

    def test_create_connection_value_error_produces_error_before_trying_again(self):
        # Arrange
        user_input = [4, 0]
        wb = load_workbook('testdata\\grahams_test_data2.xlsx')
        expected_output = "Invalid Input, please try again"

        # Act
        captured_output = io.StringIO()
        sys.stdout = captured_output
        with patch('builtins.input', side_effect=user_input):
            result = DatabaseExcel.create_connection(DatabaseExcel, wb, "")
        sys.stdout = sys.__stdout__

        # Check if the printed output includes the expected strings I'm looking for
        if expected_output in captured_output.getvalue():
            result = True
        else:
            result = False

        # Assert
        self.assertTrue(result)